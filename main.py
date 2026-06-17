from fastapi import FastAPI, HTTPException, BackgroundTasks, Header
from fastapi.staticfiles import StaticFiles
from fastapi.responses import HTMLResponse, StreamingResponse
from pydantic import BaseModel
from typing import Optional, Union
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import os
import io
import shutil
import tempfile
import threading
from datetime import date, datetime

# Carrega .env se existir
_env_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".env")
if os.path.exists(_env_path):
    for _line in open(_env_path):
        _line = _line.strip()
        if _line and not _line.startswith("#") and "=" in _line:
            _k, _v = _line.split("=", 1)
            os.environ.setdefault(_k.strip(), _v.strip())

app = FastAPI(title="Coleção de Quadrinhos")

API_KEY = os.environ.get("API_KEY", "")

def verificar_chave(x_api_key: str = Header(default="")):
    if not API_KEY:
        raise HTTPException(status_code=500, detail="API_KEY não configurada no servidor.")
    if x_api_key != API_KEY:
        raise HTTPException(status_code=401, detail="Senha incorreta.")

_BASE_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH = os.environ.get("XLSX_PATH", os.path.join(_BASE_DIR, "quadrinhos.xlsx"))
BACKUP_DIR = os.path.join(_BASE_DIR, "backups")
SHEET_NAME = "QUADRINHOS"
HEADER = ["TÍTULO", "EDIÇÃO", "EDITORA", "CATEGORIA", "LIDO", "OBS", "INCLUSAO", "VALOR", "DATA_LEITURA"]
MESES_PT = ["Janeiro","Fevereiro","Março","Abril","Maio","Junho","Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]

def _label_mes(mes_ano):
    """'05/2026' → 'Mai/2026'"""
    try:
        p = mes_ano.split("/")
        return f"{MESES_PT[int(p[0])-1]}/{p[1]}"
    except Exception:
        return mes_ano


def parse_valor(v):
    """Converte texto/numero em float (aceita '25,90', 'R$ 25,90', '25.90')."""
    if v is None or v == "":
        return None
    s = str(v).strip().replace("R$", "").replace(" ", "")
    if "," in s:
        s = s.replace(".", "").replace(",", ".")
    try:
        return round(float(s), 2)
    except ValueError:
        return None

_lock = threading.Lock()

# Cache em memória das linhas de dados, invalidado pelo mtime do arquivo.
_rows_cache = {"mtime": None, "rows": None}


def _file_mtime():
    try:
        return os.path.getmtime(XLSX_PATH) if os.path.exists(XLSX_PATH) else None
    except OSError:
        return None


def get_data_rows():
    """Retorna as linhas de dados usando cache, recarregando só quando o arquivo muda.

    Acquire o _lock internamente — não chame dentro de outro `with _lock`.
    """
    with _lock:
        mtime = _file_mtime()
        if _rows_cache["rows"] is None or _rows_cache["mtime"] != mtime:
            wb = load_wb(); ws = get_ws(wb)
            _rows_cache["rows"] = all_data_rows(ws)
            _rows_cache["mtime"] = _file_mtime()
        return _rows_cache["rows"]


def _invalidate_cache():
    _rows_cache["mtime"] = None
    _rows_cache["rows"] = None


def save_wb(wb):
    """Salva de forma atômica (arquivo temporário + replace) e gera backup diário."""
    fd, tmp = tempfile.mkstemp(suffix=".xlsx", dir=_BASE_DIR)
    os.close(fd)
    try:
        wb.save(tmp)
        os.replace(tmp, XLSX_PATH)
    except Exception:
        if os.path.exists(tmp):
            os.remove(tmp)
        raise
    _invalidate_cache()
    _backup_diario()


def _backup_diario():
    """Mantém uma cópia por dia em backups/, guardando os últimos 14 dias."""
    try:
        os.makedirs(BACKUP_DIR, exist_ok=True)
        dest = os.path.join(BACKUP_DIR, f"quadrinhos_{date.today().isoformat()}.xlsx")
        if not os.path.exists(dest):
            shutil.copy2(XLSX_PATH, dest)
            backups = sorted(f for f in os.listdir(BACKUP_DIR) if f.endswith(".xlsx"))
            for old in backups[:-14]:
                os.remove(os.path.join(BACKUP_DIR, old))
    except Exception:
        pass


def load_wb():
    if os.path.exists(XLSX_PATH):
        wb = openpyxl.load_workbook(XLSX_PATH)
        ws = get_ws(wb)
        if ws.cell(row=1, column=len(HEADER)).value != HEADER[-1]:
            c = ws.cell(row=1, column=len(HEADER), value=HEADER[-1])
            c.font = Font(bold=True, color="FFFFFF")
            c.fill = PatternFill("solid", fgColor="4F46E5")
            c.alignment = Alignment(horizontal="center")
        return wb
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    _write_header(ws)
    save_wb(wb)
    return wb


def _write_header(ws):
    for col, val in enumerate(HEADER, 1):
        cell = ws.cell(row=1, column=col, value=val)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F46E5")
        cell.alignment = Alignment(horizontal="center")


def get_ws(wb):
    if SHEET_NAME in wb.sheetnames:
        return wb[SHEET_NAME]
    return wb.active


def _edicao_key(val):
    """Ordena edições numericamente quando possível; texto ('Única') vai pro fim."""
    s = str(val).strip() if val else ""
    digits = "".join(c for c in s if c.isdigit())
    if digits:
        return (0, int(digits), s.upper())
    return (1, 0, s.upper())


def row_to_dict(row_index, row_values):
    def g(i):
        v = row_values[i] if i < len(row_values) else None
        if isinstance(v, (datetime, date)):
            return v.strftime("%d/%m/%Y")
        return str(v).strip() if v is not None else ""
    return {
        "id": row_index,
        "titulo": g(0),
        "edicao": g(1) or None,
        "editora": g(2) or None,
        "categoria": g(3) or None,
        "lido": g(4).upper() == "SIM",
        "obs": g(5) or None,
        "inclusao": g(6) or None,
        "valor": parse_valor(row_values[7]) if len(row_values) > 7 else None,
        "data_leitura": g(8) or None,
    }


def all_data_rows(ws):
    result = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(v for v in row if v):
            continue
        result.append((i, list(row)))
    return result


class Quadrinho(BaseModel):
    titulo: str
    edicao: Optional[str] = None
    editora: Optional[str] = None
    categoria: Optional[str] = None
    lido: Optional[bool] = False
    obs: Optional[str] = None
    inclusao: Optional[str] = None
    valor: Optional[Union[str, float, int]] = None
    data_leitura: Optional[str] = None


@app.get("/", response_class=HTMLResponse)
def index():
    with open(os.path.join(_BASE_DIR, "static", "index.html"), encoding="utf-8") as f:
        return f.read()


@app.get("/api/quadrinhos")
def listar(busca: str = "", categoria: str = "", editora: str = "", lido: str = "", autografado: str = "", ordenar: str = "", page: int = 1, per_page: int = 50):
    rows = get_data_rows()
    items = []
    for row_index, row_values in rows:
        d = row_to_dict(row_index, row_values)
        if busca and busca.lower() not in f"{d['titulo']} {d['editora'] or ''} {d['obs'] or ''}".lower():
            continue
        if categoria and (d.get("categoria") or "").upper() != categoria.upper():
            continue
        if editora and (d.get("editora") or "").upper() != editora.upper():
            continue
        if lido == "sim" and not d["lido"]: continue
        if lido == "nao" and d["lido"]: continue
        if autografado == "sim" and "autograf" not in (d.get("obs") or "").lower(): continue
        items.append(d)

    if ordenar == "recentes":
        def data_key(d):
            try:
                p = (d["inclusao"] or "").split("/")
                return (int(p[2]), int(p[1]), int(p[0]))
            except (IndexError, ValueError):
                return (0, 0, 0)
        items.sort(key=data_key, reverse=True)
    elif ordenar == "nao_lidos":
        items.sort(key=lambda d: (d["lido"], (d["titulo"] or "").upper(), _edicao_key(d["edicao"])))
    else:
        items.sort(key=lambda d: ((d["titulo"] or "").upper(), _edicao_key(d["edicao"])))

    total = len(items)
    start = (page - 1) * per_page
    return {"total": total, "page": page, "per_page": per_page, "items": items[start:start + per_page]}


@app.get("/api/quadrinhos/{row_index}")
def obter(row_index: int):
    for i, r in get_data_rows():
        if i == row_index:
            return row_to_dict(i, r)
    raise HTTPException(404, "Não encontrado")


@app.post("/api/quadrinhos", status_code=201)
def criar(q: Quadrinho, background: BackgroundTasks, x_api_key: str = Header(default="")):
    verificar_chave(x_api_key)
    # Auto-preenche DATA_LEITURA com mês atual se marcado como lido
    dl = q.data_leitura or (date.today().strftime("%m/%Y") if q.lido else "")
    valores = [q.titulo, q.edicao or "", q.editora or "", q.categoria or "",
               "SIM" if q.lido else "NÃO", q.obs or "",
               q.inclusao or date.today().strftime("%d/%m/%Y"),
               parse_valor(q.valor), dl]
    with _lock:
        wb = load_wb(); ws = get_ws(wb)
        ws.append(valores)
        save_wb(wb)
    background.add_task(_sync_to_sheets, "append", None, valores)
    return {"ok": True}


@app.put("/api/quadrinhos/{row_index}")
def atualizar(row_index: int, q: Quadrinho, background: BackgroundTasks, x_api_key: str = Header(default="")):
    verificar_chave(x_api_key)
    # Auto-preenche DATA_LEITURA com mês atual ao marcar como lido (se ainda não tiver)
    dl = q.data_leitura or ""
    if q.lido and not dl:
        # Verifica se já tinha DATA_LEITURA antes
        atual = next((r for i, r in get_data_rows() if i == row_index), None)
        dl_atual = str(atual[8]).strip() if atual and len(atual) > 8 and atual[8] else ""
        dl = dl_atual or date.today().strftime("%m/%Y")
    valores = [q.titulo, q.edicao or "", q.editora or "", q.categoria or "",
               "SIM" if q.lido else "NÃO", q.obs or "", q.inclusao or "",
               parse_valor(q.valor), dl]
    with _lock:
        wb = load_wb(); ws = get_ws(wb)
        if row_index < 2 or row_index > ws.max_row:
            raise HTTPException(404, "Não encontrado")
        for col, val in enumerate(valores, 1):
            ws.cell(row=row_index, column=col, value=val)
        save_wb(wb)
    background.add_task(_sync_to_sheets, "update", row_index, valores)
    return {"ok": True}


@app.delete("/api/quadrinhos/{row_index}")
def deletar(row_index: int, background: BackgroundTasks, x_api_key: str = Header(default="")):
    verificar_chave(x_api_key)
    with _lock:
        wb = load_wb(); ws = get_ws(wb)
        if row_index < 2 or row_index > ws.max_row:
            raise HTTPException(404, "Não encontrado")
        ws.delete_rows(row_index)
        save_wb(wb)
    background.add_task(_sync_to_sheets, "delete", row_index, None)
    return {"ok": True}


@app.get("/api/verificar")
def verificar(x_api_key: str = Header(default="")):
    verificar_chave(x_api_key)
    return {"ok": True}


@app.get("/api/stats")
def stats():
    rows = get_data_rows()
    total = len(rows)
    lidos = sum(1 for _, r in rows if len(r) > 4 and str(r[4]).strip().upper() == "SIM")
    return {"total": total, "lidos": lidos, "nao_lidos": total - lidos}


@app.get("/api/filtros")
def filtros():
    rows = get_data_rows()
    cat_count: dict = {}
    ed_count: dict = {}
    for _, r in rows:
        if len(r) > 3 and r[3]:
            c = str(r[3]).strip(); cat_count[c] = cat_count.get(c, 0) + 1
        if len(r) > 2 and r[2]:
            e = str(r[2]).strip(); ed_count[e] = ed_count.get(e, 0) + 1
    categorias = [{"nome": k, "count": v} for k, v in sorted(cat_count.items())]
    editoras = [{"nome": k, "count": v} for k, v in sorted(ed_count.items())]
    return {"categorias": categorias, "editoras": editoras}


@app.get("/api/sugestoes")
def sugestoes():
    rows = get_data_rows()
    titulos = sorted(set(str(r[0]).strip() for _, r in rows if r[0]))
    editoras = sorted(set(str(r[2]).strip() for _, r in rows if len(r) > 2 and r[2]))
    categorias = sorted(set(str(r[3]).strip() for _, r in rows if len(r) > 3 and r[3]))
    return {"titulos": titulos, "editoras": editoras, "categorias": categorias}


def _http_json(url, timeout=6, headers=None):
    """GET simples que devolve JSON (ou None em qualquer falha)."""
    import urllib.request
    import json as _json
    try:
        req = urllib.request.Request(url, headers=headers or {"User-Agent": "quadrinhos-app/1.0"})
        with urllib.request.urlopen(req, timeout=timeout) as resp:
            return _json.loads(resp.read().decode("utf-8"))
    except Exception:
        return None


@app.get("/api/isbn/{isbn}")
def buscar_isbn(isbn: str):
    """Consulta dados de um livro por ISBN em fontes gratuitas.
    Ordem: Mercado Editorial (base brasileira/CBL) -> Google Books -> OpenLibrary."""
    limpo = "".join(c for c in isbn if c.isdigit() or c in "Xx")
    if not limpo:
        raise HTTPException(status_code=400, detail="ISBN inválido")

    # 1) Mercado Editorial (Agência Brasileira do ISBN / dados CBL)
    d = _http_json(f"https://api.mercadoeditorial.org/api/v1.2/book?isbn={limpo}")
    if d and d.get("books"):
        b = d["books"][0]
        titulo = (b.get("title") or "").strip()
        if titulo:
            editora = ""
            pub = b.get("publishers") or b.get("publisher")
            if isinstance(pub, list) and pub:
                editora = (pub[0].get("name") if isinstance(pub[0], dict) else str(pub[0])) or ""
            elif isinstance(pub, dict):
                editora = pub.get("name", "")
            elif isinstance(pub, str):
                editora = pub
            return {"fonte": "mercadoeditorial", "titulo": titulo,
                    "editora": editora.strip(), "subtitulo": (b.get("subtitle") or "").strip()}

    # 2) Google Books
    d = _http_json(f"https://www.googleapis.com/books/v1/volumes?q=isbn:{limpo}")
    if d and d.get("items"):
        info = d["items"][0].get("volumeInfo", {})
        if info.get("title"):
            return {"fonte": "googlebooks", "titulo": info.get("title", ""),
                    "editora": info.get("publisher", ""), "subtitulo": info.get("subtitle", "")}

    # 3) OpenLibrary
    d = _http_json(f"https://openlibrary.org/api/books?bibkeys=ISBN:{limpo}&format=json&jscmd=data")
    if d and d.get(f"ISBN:{limpo}"):
        obj = d[f"ISBN:{limpo}"]
        pubs = obj.get("publishers") or []
        editora = pubs[0].get("name", "") if pubs and isinstance(pubs[0], dict) else ""
        if obj.get("title"):
            return {"fonte": "openlibrary", "titulo": obj.get("title", ""),
                    "editora": editora, "subtitulo": obj.get("subtitle", "")}

    return {"fonte": None, "titulo": "", "editora": "", "subtitulo": ""}


def _inc_to_str(v) -> str:
    """Normaliza valor de célula inclusao para 'dd/mm/yyyy', seja datetime ou string."""
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%d/%m/%Y")
    return str(v).strip()


def _periodo_key(inc):
    """Retorna 'YYYY-MM' a partir de datetime ou string 'dd/mm/yyyy'."""
    try:
        s = _inc_to_str(inc)
        if not s:
            return None
        parts = s.split("/")
        key = f"{parts[2]}-{parts[1].zfill(2)}"
        return key if len(key) == 7 else None
    except Exception:
        return None


@app.get("/api/dashboard")
def dashboard(editora: str = "", categoria: str = "", status: str = "", periodo: str = ""):
    rows = get_data_rows()
    data = rows
    if editora and editora != "TODOS":
        data = [(i, r) for i, r in data if len(r) > 2 and str(r[2]).strip() == editora]
    if categoria and categoria != "TODOS":
        data = [(i, r) for i, r in data if len(r) > 3 and str(r[3]).strip() == categoria]
    if status == "SIM":
        data = [(i, r) for i, r in data if len(r) > 4 and str(r[4]).strip().upper() == "SIM"]
    elif status == "NÃO":
        data = [(i, r) for i, r in data if len(r) > 4 and str(r[4]).strip().upper() == "NÃO"]

    meses_disponiveis = sorted(
        {k for _, r in data if len(r) > 6 and r[6] and (k := _periodo_key(r[6]))},
        reverse=True,
    )
    if periodo and periodo != "TODOS":
        data = [(i, r) for i, r in data if len(r) > 6 and r[6] and _periodo_key(str(r[6]).strip()) == periodo]

    total = len(data)
    lidos = sum(1 for _, r in data if len(r) > 4 and str(r[4]).strip().upper() == "SIM")
    cat_count: dict = {}
    ed_count: dict = {}
    evolucao: dict = {}
    gasto_mensal: dict = {}
    total_gasto = 0.0

    for _, r in data:
        cat = str(r[3]).strip() if len(r) > 3 and r[3] else ""
        ed = str(r[2]).strip() if len(r) > 2 and r[2] else ""
        inc = _inc_to_str(r[6]) if len(r) > 6 else ""
        valor = parse_valor(r[7]) if len(r) > 7 else None
        if cat: cat_count[cat] = cat_count.get(cat, 0) + 1
        if ed: ed_count[ed] = ed_count.get(ed, 0) + 1
        if valor: total_gasto += valor
        if inc:
            try:
                parts = inc.split("/")
                key = f"{parts[2]}-{parts[1].zfill(2)}"
                if len(key) == 7:
                    evolucao[key] = evolucao.get(key, 0) + 1
                    if valor: gasto_mensal[key] = round(gasto_mensal.get(key, 0) + valor, 2)
            except: pass

    ed_sorted = sorted(ed_count.items(), key=lambda x: -x[1])
    cat_sorted = sorted(cat_count.items(), key=lambda x: -x[1])

    # Insights de gasto
    gasto_editora: dict = {}
    itens_com_valor = 0
    for _, r in data:
        valor = parse_valor(r[7]) if len(r) > 7 else None
        if valor:
            itens_com_valor += 1
            ed = str(r[2]).strip() if len(r) > 2 and r[2] else "(sem editora)"
            gasto_editora[ed] = round(gasto_editora.get(ed, 0) + valor, 2)
    gasto_medio = round(total_gasto / itens_com_valor, 2) if itens_com_valor else 0
    top_editora_gasto = max(gasto_editora.items(), key=lambda x: x[1]) if gasto_editora else None
    mes_maior_gasto = max(gasto_mensal.items(), key=lambda x: x[1]) if gasto_mensal else None

    return {
        "total": total, "lidos": lidos, "nao_lidos": total - lidos,
        "percentual": lidos / total if total else 0,
        "categorias": cat_count, "editoras": ed_count, "evolucao": evolucao,
        "gasto_mensal": gasto_mensal, "total_gasto": round(total_gasto, 2),
        "meses_disponiveis": meses_disponiveis,
        "editoras_sorted": [[k, v] for k, v in ed_sorted],
        "categorias_sorted": [[k, v] for k, v in cat_sorted],
        "insights": {
            "top_editora": ed_sorted[0][0] if ed_sorted else "-",
            "top_categoria": cat_sorted[0][0] if cat_sorted else "-",
        },
        "gasto_insights": {
            "gasto_medio": gasto_medio,
            "itens_com_valor": itens_com_valor,
            "top_editora_gasto": ({"nome": top_editora_gasto[0], "valor": top_editora_gasto[1]}
                                  if top_editora_gasto else None),
            "mes_maior_gasto": ({"mes": mes_maior_gasto[0], "valor": mes_maior_gasto[1]}
                                if mes_maior_gasto else None),
        },
    }


@app.get("/api/sugerir")
def sugerir():
    import random
    rows = get_data_rows()
    nao_lidos = [r for _, r in rows if len(r) > 4 and str(r[4]).strip().upper() == "NÃO"]
    if not nao_lidos:
        return {"titulo": "Você já leu tudo! 😎", "edicao": ""}
    s = random.choice(nao_lidos)
    return {"titulo": str(s[0]).strip(), "edicao": str(s[1]).strip() if len(s) > 1 and s[1] else ""}


@app.get("/api/exportar")
def exportar():
    rows = get_data_rows()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QUADRINHOS"
    widths = [50, 12, 20, 12, 8, 25, 14, 12, 14]
    for col, (h, w) in enumerate(zip(HEADER, widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="4F46E5")
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[get_column_letter(col)].width = w
    for row_i, (_, r) in enumerate(rows, 2):
        for col, val in enumerate(r[:9], 1):
            ws.cell(row=row_i, column=col, value=val)
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=quadrinhos.xlsx"},
    )


CREDENTIALS_PATH = os.path.join(_BASE_DIR, "google_credentials.json")
GOOGLE_SHEETS_ID = os.environ.get("GOOGLE_SHEETS_ID", "")
GS_SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]

# Cache da worksheet do Google Sheets (autoriza uma vez e reaproveita).
_gs_lock = threading.Lock()
_gs_cache = {"ws": None}


def _load_gs_credentials():
    """Carrega as credenciais: prioriza a env GOOGLE_CREDENTIALS_JSON (útil no Render),
    senão usa o arquivo google_credentials.json local."""
    from google.oauth2.service_account import Credentials
    raw = os.environ.get("GOOGLE_CREDENTIALS_JSON")
    if raw:
        import json
        return Credentials.from_service_account_info(json.loads(raw), scopes=GS_SCOPES)
    if os.path.exists(CREDENTIALS_PATH):
        return Credentials.from_service_account_file(CREDENTIALS_PATH, scopes=GS_SCOPES)
    raise HTTPException(
        status_code=503,
        detail="Credenciais do Google ausentes (GOOGLE_CREDENTIALS_JSON ou google_credentials.json).",
    )


def _gsheet_ws(force=False):
    """Retorna a worksheet 'QUADRINHOS' (cacheada). Levanta HTTPException se mal configurado."""
    if _gs_cache["ws"] is not None and not force:
        return _gs_cache["ws"]
    if not GOOGLE_SHEETS_ID:
        raise HTTPException(status_code=503, detail="GOOGLE_SHEETS_ID não configurado.")
    try:
        import gspread
    except ImportError:
        raise HTTPException(status_code=503, detail="Instale: pip install gspread google-auth")
    creds = _load_gs_credentials()
    try:
        gc = gspread.authorize(creds)
        sh = gc.open_by_key(GOOGLE_SHEETS_ID)
        try:
            ws = sh.worksheet("QUADRINHOS")
        except gspread.WorksheetNotFound:
            ws = sh.add_worksheet(title="QUADRINHOS", rows=2000, cols=10)
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao conectar ao Google Sheets: {e}")
    _gs_cache["ws"] = ws
    return ws


def _cell_to_str(v):
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return v.strftime("%d/%m/%Y")
    return str(v).strip()


def _row_for_sheets(values):
    return [_cell_to_str(values[i]) if i < len(values) else "" for i in range(9)]


def _push_all_to_sheets():
    """Reescreve TODO o Google Sheets a partir do Excel local (best-effort, background).

    Em vez de sincronizar por índice de linha (frágil: deletes deslocam as linhas e
    qualquer falha best-effort dessincroniza Sheets×Excel), reescreve tudo de uma vez.
    Assim o Sheets é sempre um espelho fiel do Excel após cada alteração.
    """
    try:
        ws_gs = _gsheet_ws()
        rows = get_data_rows()
        sheet_data = [HEADER]
        for _, r in rows:
            sheet_data.append(_row_for_sheets(r))
        with _gs_lock:
            ws_gs.clear()
            ws_gs.update(sheet_data, value_input_option="USER_ENTERED")
    except Exception as e:
        print(f"[sheets-sync] falha no push completo: {e}")


def _sync_to_sheets(action=None, row_index=None, values=None):
    """Compat: qualquer alteração dispara um push completo (evita drift por índice)."""
    _push_all_to_sheets()


def _pull_from_sheets():
    """Baixa todos os dados do Google Sheets e reescreve o Excel local. Retorna a contagem."""
    ws_gs = _gsheet_ws()
    with _gs_lock:
        valores = ws_gs.get_all_values()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    _write_header(ws)
    n = 0
    for r in valores[1:]:  # pula o cabeçalho
        if not any(str(c).strip() for c in r):
            continue
        ws.append([(r[i] if i < len(r) else "") for i in range(9)])
        n += 1
    with _lock:
        save_wb(wb)
    return n


@app.get("/api/sincronizar-sheets")
def sincronizar_sheets(x_api_key: str = Header(default="")):
    verificar_chave(x_api_key)
    """Envia TODOS os dados do Excel para o Google Sheets (Excel → Sheets)."""
    ws_gs = _gsheet_ws()
    rows = get_data_rows()
    sheet_data = [HEADER]
    for _, r in rows:
        sheet_data.append(_row_for_sheets(r))
    try:
        with _gs_lock:
            ws_gs.clear()
            ws_gs.update(sheet_data, value_input_option="USER_ENTERED")
            ws_gs.format("A1:I1", {
                "backgroundColor": {"red": 0.26, "green": 0.22, "blue": 0.79},
                "textFormat": {"bold": True, "foregroundColor": {"red": 1, "green": 1, "blue": 1}},
                "horizontalAlignment": "CENTER",
            })
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao escrever no Sheets: {e}")
    return {"ok": True, "total": len(rows),
            "mensagem": f"{len(rows)} itens enviados ao Google Sheets."}


@app.get("/api/puxar-sheets")
def puxar_sheets(x_api_key: str = Header(default="")):
    verificar_chave(x_api_key)
    """Baixa TODOS os dados do Google Sheets para o app (Sheets → Excel)."""
    try:
        n = _pull_from_sheets()
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Erro ao puxar do Sheets: {e}")
    return {"ok": True, "total": n,
            "mensagem": f"{n} itens puxados do Google Sheets."}


@app.on_event("startup")
def _startup_pull():
    """No Render (disco efêmero), puxa os dados do Sheets ao iniciar.
    Ative com a env PULL_ON_STARTUP=true. Localmente fica desligado por padrão."""
    if os.environ.get("PULL_ON_STARTUP", "").strip().lower() in ("1", "true", "yes"):
        try:
            n = _pull_from_sheets()
            print(f"[startup] {n} itens carregados do Google Sheets.")
        except Exception as e:
            print(f"[startup] pull do Sheets falhou (usando arquivo local): {e}")


@app.get("/api/historico")
def historico(ano: str = ""):
    rows = get_data_rows()
    hoje = date.today()
    ano_filtro = ano or str(hoje.year)
    mes_atual = f"{hoje.month:02d}/{hoje.year}"
    mesmo_mes_ano_ant = f"{hoje.month:02d}/{hoje.year - 1}"

    por_mes: dict = {}
    cat_count_ano: dict = {}
    total_ano = 0
    # Contagens do comparativo (independentes do ano filtrado)
    mes_atual_count = 0
    mes_ant_count = 0

    for row_i, r in rows:
        dl = str(r[8]).strip() if len(r) > 8 and r[8] else ""
        if not dl or "/" not in dl:
            continue
        partes = dl.split("/")
        if len(partes) != 2:
            continue

        # Comparativo mês atual × mesmo mês ano anterior (qualquer ano filtrado)
        if dl == mes_atual:
            mes_atual_count += 1
        elif dl == mesmo_mes_ano_ant:
            mes_ant_count += 1

        ano_dl = partes[1]
        if ano_dl != ano_filtro:
            continue

        d = row_to_dict(row_i, r)
        total_ano += 1
        cat = d.get("categoria") or "Outro"
        cat_count_ano[cat] = cat_count_ano.get(cat, 0) + 1
        por_mes.setdefault(dl, []).append(d)

    # Ordena meses desc (mais recente primeiro)
    def _mes_key(m):
        p = m.split("/")
        return (int(p[1]), int(p[0]))

    meses_ord = sorted(por_mes.keys(), key=_mes_key, reverse=True)

    cat_fav = max(cat_count_ano.items(), key=lambda x: x[1]) if cat_count_ano else None

    anos = sorted(
        {str(r[8]).split("/")[1] for _, r in rows
         if len(r) > 8 and r[8] and "/" in str(r[8])},
        reverse=True,
    )

    return {
        "stats": {
            "ano": ano_filtro,
            "total_ano": total_ano,
            "mes_atual": mes_atual,
            "mes_atual_label": _label_mes(mes_atual),
            "mes_atual_count": mes_atual_count,
            "mesmo_mes_ano_ant": mesmo_mes_ano_ant,
            "mesmo_mes_ano_ant_label": _label_mes(mesmo_mes_ano_ant),
            "mesmo_mes_ano_ant_count": mes_ant_count,
            "categoria_favorita": cat_fav[0] if cat_fav else None,
            "categoria_favorita_count": cat_fav[1] if cat_fav else 0,
            "categoria_favorita_pct": round(cat_fav[1] / total_ano * 100) if cat_fav and total_ano else 0,
        },
        "por_mes": [
            {
                "mes_ano": m,
                "label": _label_mes(m),
                "count": len(por_mes[m]),
                "items": sorted(por_mes[m], key=lambda x: (x["titulo"], x["edicao"] or "")),
            }
            for m in meses_ord
        ],
        "anos_disponiveis": anos,
    }


app.mount("/static", StaticFiles(directory=os.path.join(_BASE_DIR, "static")), name="static")
