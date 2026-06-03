#!/usr/bin/env python3
"""
Sincroniza dados do Google Drive (Excel ou Sheets)
Uso: python3 sync_from_drive.py <FILE_ID_ou_LINK>
"""

import sys
import urllib.request
import csv
import openpyxl
from openpyxl.styles import Font, PatternFill
from datetime import datetime

def get_file_id(arg):
    """Extrai ID do arquivo de um link ou usa diretamente"""
    if "drive.google.com" in arg:
        # Extrai ID de link de compartilhamento
        if "/d/" in arg:
            return arg.split("/d/")[1].split("/")[0]
        if "id=" in arg:
            return arg.split("id=")[1].split("&")[0]
    return arg.strip()

def download_from_drive(file_id):
    """Baixa arquivo do Google Drive"""
    url = f"https://drive.google.com/uc?export=download&id={file_id}"
    print(f"📥 Baixando arquivo {file_id[:20]}... do Google Drive")

    try:
        with urllib.request.urlopen(url, timeout=15) as response:
            return response.read()
    except Exception as e:
        print(f"❌ Erro ao baixar: {e}")
        return None

def excel_to_csv_data(excel_bytes):
    """Converte Excel em CSV data (retorna linhas)"""
    import io
    try:
        wb = openpyxl.load_workbook(io.BytesIO(excel_bytes))
        ws = wb.active

        rows = []
        for row in ws.iter_rows(values_only=True):
            rows.append([str(v) if v is not None else "" for v in row])

        return rows
    except Exception as e:
        print(f"❌ Erro ao ler Excel: {e}")
        return None

def sync_to_local(rows):
    """Escreve dados ao quadrinhos.xlsx local"""
    if not rows or len(rows) < 2:
        print("❌ Nenhum dado para sincronizar")
        return False

    print(f"📊 Processando {len(rows)-1} itens...")

    # Cria workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "QUADRINHOS"

    # Header
    HEADER = ["TÍTULO", "EDIÇÃO", "EDITORA", "CATEGORIA", "LIDO", "OBS", "INCLUSAO", "VALOR"]
    ws.append(HEADER)

    # Formata header
    header_fill = PatternFill(start_color="4338CA", end_color="4338CA", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font

    # Escreve dados (assume primeiras 8 colunas)
    added = 0
    for row_data in rows[1:]:
        if not row_data or not str(row_data[0]).strip():
            continue

        new_row = []
        for i in range(8):
            val = row_data[i] if i < len(row_data) else ""
            if isinstance(val, str):
                val = val.strip()
            new_row.append(val if val else None)

        ws.append(new_row)
        added += 1

    # Ajusta colunas
    widths = [40, 12, 20, 12, 8, 25, 14, 12]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[chr(64+i)].width = w

    # Salva
    ws.title = "QUADRINHOS"
    wb.save("quadrinhos.xlsx")

    print(f"✅ Sincronizado: {added} itens salvos em quadrinhos.xlsx")
    print(f"⏰ {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}")
    return True

def main():
    if len(sys.argv) < 2:
        print("Uso: python3 sync_from_drive.py <FILE_ID_ou_LINK>")
        print("\nExemplo:")
        print("  python3 sync_from_drive.py 1Bd-ssCIUvnep21dwZY5ADCsVn0mwXIBs07PC5ncNLbI")
        print("  python3 sync_from_drive.py 'https://drive.google.com/file/d/1Bd-ssCIUvnep21dwZY5ADCsVn0mwXIBs07PC5ncNLbI/view'")
        sys.exit(1)

    file_id = get_file_id(sys.argv[1])
    print(f"🔍 ID do arquivo: {file_id[:20]}...")

    # Baixa
    excel_bytes = download_from_drive(file_id)
    if not excel_bytes:
        sys.exit(1)

    # Converte
    rows = excel_to_csv_data(excel_bytes)
    if not rows:
        sys.exit(1)

    # Sincroniza
    if sync_to_local(rows):
        print("\n🎉 Pronto! Reinicie o servidor para ver os dados atualizados.")
        sys.exit(0)
    else:
        sys.exit(1)

if __name__ == "__main__":
    main()
