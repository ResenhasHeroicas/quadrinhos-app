#!/usr/bin/env python3
"""
Importa histórico de leitura das notas para quadrinhos.xlsx.
Define DATA_LEITURA (MM/YYYY) nos itens encontrados na coleção.
"""

import os, re, unicodedata, openpyxl
from difflib import SequenceMatcher
from collections import defaultdict
from openpyxl.styles import Font, PatternFill, Alignment

# ── Notas de leitura ───────────────────────────────────────────────────────────
NOTAS = """
📚Quadrinhos 2025

✅ Janeiro
* Dragonball 7,8,9,10,11,12,13,14,15,16
* Battle Royalle 2
* Crazy Food Truck 1,2

✅ Fevereiro
* Battle Royale 3
* One-Punch Man 17, 18, 19

✅ Março
* One-Punch Man 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30, 31
* A Saga do Homem-Aranha 1, 2
* DandaDan 1, 2, 3, 4
* Battle Royale 4

✅ Abril
* DandaDan 5, 6, 7, 8, 9, 10, 11, 12, 13
* 20th Century Boys 1, 2
* A Saga do Homem-Aranha 3
* Tartarugas Ninja Clássica 1, 2
* Sawala

✅ Maio
* Tartarugas Ninja Clássica 3, 4, 5, 6
* Os Guardiões do Maser 1
* O Último Ronin
* A Saga do Homem-Aranha 4, 5
* DandaDan 14
* Só as Melhores Paulo Moreira

✅ Junho
* Akira 1, 2, 3

✅ Julho
* Battle Royale 5
* Sense Life 1
* Lobo solitário 1, 2
* Kagurabachi 1
* Dandadan 15
* Akira 4

✅ Agosto
* Tartarugas Ninja IDW 1
* Shigurui 1
* 20th Century Boys 3
* Wolverine (2025) 1
* One Punch Man 32

✅ Setembro
* Wolverine (2025) 2
* Dandadan 16
* Asterix Omnibus 1
* Usagi Yojimbo 4
* Absolute Batman 1

✅ Outubro
* Usagi Yojimbo 5, 6
* O Vampiro que ri 1
* 20th Century Boys 4
* Shigurui 2
* Ultimate Homem-Aranha 2
* Frankenstein
* Vampiros
* Hajime no Ippo 1
* Wolverine (2025) 3
* Contagem de Corpos

✅ Novembro
* Tartarugas Ninja IDW 2
* One Punch Man 33
* Thorgal Clássico 1
* Berserk 1, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 14, 15
* Souichi
* Kagurabachi 2
* Batman 1
* Homem aranha (2025) 1
* Wolverine (2025) 4
* Absolute Superman 1
* Aguardela
* Asterix Omnibus 2

✅ Dezembro
* Berserk 16
* Popeye Um Homem ao Mar
* Shigurui 3
* Kagurabachi 3, 4
* 20th Century Boys 5
* Absolute Mulher-Maravilha 1
* Deadpool e Batman
* Batman e Deadpool
* Homem aranha (2025) 2

📚Quadrinhos 2026

✅ Janeiro
* Batman Ano 100 - 1 e 2
* Absolute Batman 2
* A lenda de Musashi - Goseki Kojima
* Berserk 17, 18, 19, 20, 21, 22, 23, 24, 25, 26, 27, 28, 29, 30
* Batman (2025) 2, 3
* Absolute Superman 2
* Dandadan 17, 18
* One Piece Episode A 1

✅ Fevereiro
* Shigurui 4
* Hajime no Ippo 2, 3
* O Hobbit em Quadrinhos
* O Espetacular Homem-Aranha (2025) - 3
* Absolute Caçador de Marte 1
* One Punch Man 34
* O Eternauta 1969
* Os Smurfs Integral 1

✅ Março
* Absolute Mulher Maravilha 2
* Undertaker Edição Definitiva
* Absolute Batman 3
* Batman (2025) 4, 5
* O Espetacular Homem-Aranha (2025) - 4, 5, 6
* Kagurabachi 5
* 20th Century boys Perfect Edition 6
* Absolute Superman 3

✅ Abril
* Paraíso - O Vampiro que Ri 2
* Astro Boy Big 1
* One Piece Episode A 2
* Absolute Mulher Maravilha 3
* Shigurui 5
* Absolute Caçador de Marte 2
* Berserk 31, 32, 33, 34, 35, 36
* Hajime no Ippo 4
* Batman (2025) 6
* Transformers (2023) 1
* 20th Century Boys Perfect Edition 7

✅ Maio
* Blade - A Lâmina do Imortal Big 1, 2, 3
* Absolute Batman 4
* Elric - O Campeão Eterno
* Dandadan 19, 20
* Kagurabachi 6
* Batman (2025) 7
* Chico Bento (2026) 1
* Superman Vs Homem-Aranha (1976)
* One Punch Man 35
"""

# ── Constantes ─────────────────────────────────────────────────────────────────
MESES = {
    'janeiro':'01','fevereiro':'02','março':'03','abril':'04',
    'maio':'05','junho':'06','julho':'07','agosto':'08',
    'setembro':'09','outubro':'10','novembro':'11','dezembro':'12'
}
XLSX_PATH  = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'quadrinhos.xlsx')
COL_TITULO = 1
COL_EDICAO = 2
COL_DATA   = 9   # nova coluna DATA_LEITURA

# Aliases: como aparece na nota → título exato na coleção (normalizado)
ALIASES = {
    'shigurui':                          'shigurui frenesi da morte',
    'wolverine 2025':                    'wolverine quinta serie',
    '20th century boys':                 '20th century boys edicao definitiva',
    '20th century boys perfect edition': '20th century boys edicao definitiva',
    '20th century boys edicao definitiva': '20th century boys edicao definitiva',
    'souichi':                           'as egocentricas maldicoes de souichi',
    'aguardela':                         '100 discos para conhecer aguardela',
    'contagem de corpos':                'contagem de corpos tartarugas ninja',
    'batman 2025':                       'batman 2025',
    'batman':                            'batman 2025',
    'homem aranha 2025':                 'o espetacular homem aranha quinta serie',
    'paraiso o vampiro que ri':          'o vampiro que ri paraiso',
}


# ── Helpers ────────────────────────────────────────────────────────────────────
def norm(s):
    s = str(s).lower()
    s = unicodedata.normalize('NFD', s)
    s = ''.join(c for c in s if unicodedata.category(c) != 'Mn')
    s = re.sub(r'[^\w\s]', ' ', s)
    return re.sub(r'\s+', ' ', s).strip()

def sim(a, b):
    base = SequenceMatcher(None, norm(a), norm(b)).ratio()
    # Bônus: se todas as palavras de 'a' aparecem em 'b' (ex: "Shigurui" em "Shigurui Frenesi da Morte")
    palavras_a = set(norm(a).split())
    palavras_b = set(norm(b).split())
    if palavras_a and palavras_a.issubset(palavras_b):
        base = max(base, 0.80)
    return base

def eh_ano(n):
    """Retorna True se n for um número de 4 dígitos parecido com ano (1800-2099)."""
    return bool(re.match(r'^(1[89]\d\d|20\d\d)$', n))

def extrair_edicoes(texto):
    texto = re.sub(r'\be\b', ',', texto)
    partes = [p.strip() for p in texto.split(',')]
    # Ignora anos (ex: "1969", "2025") como edições
    return [p for p in partes if re.match(r'^\d+$', p) and not eh_ano(p)]

def parsear_linha(linha):
    """
    Retorna (titulo, [edicoes]).
    Casos tratados:
      "Batman Ano 100 - 1 e 2"              → ("Batman Ano 100", ["1","2"])
      "Blade - A Lâmina do Imortal Big 1,2" → ("Blade - A Lâmina do Imortal Big", ["1","2"])
      "A lenda de Musashi - Goseki Kojima"  → ("A lenda de Musashi - Goseki Kojima", [])
      "One Punch Man 17, 18, 19"            → ("One Punch Man", ["17","18","19"])
      "O Hobbit em Quadrinhos"              → ("O Hobbit em Quadrinhos", [])
      "O Eternauta 1969"                    → ("O Eternauta 1969", [])  ← ano = parte do título
    """
    linha = linha.strip()

    # Padrão 1: " - " onde o lado direito é SOMENTE números/vírgulas (não anos)
    m = re.match(r'^(.+?)\s*-\s*([\d][\d,\s]*(?:\be\b\s*\d+)?)$', linha)
    if m:
        edicoes = extrair_edicoes(m.group(2))
        if edicoes:
            return m.group(1).strip(), edicoes

    # Padrão 2: números no final da linha (ignora anos de 4 dígitos)
    m = re.match(r'^(.+?)\s+((?:\d+)(?:\s*,\s*\d+)*)$', linha)
    if m:
        edicoes = [e.strip() for e in m.group(2).split(',') if not eh_ano(e.strip())]
        if edicoes:
            return m.group(1).strip(), edicoes

    # Padrão 3: sem edição (ou título termina com ano)
    return linha, []

def parsear_notas(texto):
    """Retorna lista de (mes_ano, titulo, edicao_ou_None)."""
    itens = []
    ano_atual = mes_atual = None

    for linha in texto.splitlines():
        linha = linha.strip()

        # Cabeçalho de ano
        if ('quadrinho' in linha.lower() or '📚' in linha):
            m = re.search(r'\b(20\d\d)\b', linha)
            if m:
                ano_atual = m.group(1)
                mes_atual = None
            continue

        # Cabeçalho de mês
        if '✅' in linha:
            ln = linha.lower()
            for nome, num in MESES.items():
                if nome in ln:
                    mes_atual = num
                    break
            continue

        # Item de leitura
        if linha.startswith('*') and mes_atual and ano_atual:
            conteudo = linha.lstrip('*').strip()
            if not conteudo:
                continue
            titulo, edicoes = parsear_linha(conteudo)
            mes_ano = f"{mes_atual}/{ano_atual}"
            if edicoes:
                for ed in edicoes:
                    itens.append((mes_ano, titulo, ed))
            else:
                itens.append((mes_ano, titulo, None))

    return itens


# ── Main ───────────────────────────────────────────────────────────────────────
def main():
    print("📚 Importador de Histórico de Leitura")
    print("=" * 55)

    wb = openpyxl.load_workbook(XLSX_PATH)
    ws = wb.active

    # Adiciona coluna DATA_LEITURA se ainda não existir
    cabecalho = [ws.cell(row=1, column=c).value for c in range(1, 12)]
    if 'DATA_LEITURA' not in cabecalho:
        cell = ws.cell(row=1, column=COL_DATA, value='DATA_LEITURA')
        cell.font  = Font(bold=True, color='FFFFFF')
        cell.fill  = PatternFill('solid', fgColor='4F46E5')
        cell.alignment = Alignment(horizontal='center')
        print("✅ Coluna DATA_LEITURA criada\n")
    else:
        print("✅ Coluna DATA_LEITURA já existia\n")

    # Indexa coleção por título normalizado
    colecao = []
    by_title = defaultdict(list)
    for row_i in range(2, ws.max_row + 1):
        titulo = ws.cell(row=row_i, column=COL_TITULO).value
        edicao = ws.cell(row=row_i, column=COL_EDICAO).value
        if titulo:
            t = str(titulo).strip()
            e = str(edicao).strip() if edicao else ''
            colecao.append((row_i, t, e))
            by_title[norm(t)].append((row_i, t, e))

    print(f"📦 {len(colecao)} itens na coleção")

    itens_nota = parsear_notas(NOTAS)
    print(f"📋 {len(itens_nota)} entradas nas notas\n")

    matched    = 0
    conflitos  = 0
    nao_achado = []

    for mes_ano, titulo_nota, edicao_nota in itens_nota:

        # Verifica alias primeiro
        norm_nota = norm(titulo_nota)
        melhor_key = ALIASES.get(norm_nota)
        if melhor_key and melhor_key in by_title:
            melhor_score = 1.0
        else:
            melhor_key   = None
            melhor_score = 0
            for key in by_title:
                s = sim(norm_nota, key)
                if s > melhor_score:
                    melhor_score = s
                    melhor_key   = key

        LIMIAR = 0.70
        if melhor_score < LIMIAR or melhor_key is None:
            nao_achado.append((mes_ano, titulo_nota, edicao_nota, melhor_score))
            continue

        candidatos = by_title[melhor_key]

        # Filtra por edição quando fornecida
        if edicao_nota:
            matches = [c for c in candidatos
                       if norm(c[2]) == norm(edicao_nota) or c[2].strip() == edicao_nota.strip()]
            if not matches:
                nao_achado.append((mes_ano, titulo_nota, edicao_nota, melhor_score))
                continue
        else:
            matches = candidatos

        for row_i, titulo_col, edicao_col in matches:
            atual = ws.cell(row=row_i, column=COL_DATA).value
            if atual and atual != mes_ano:
                conflitos += 1
                ed_str = f" #{edicao_col}" if edicao_col else ""
                print(f"  ⚠️  Conflito [{titulo_col}{ed_str}]: "
                      f"já tem '{atual}', nota diz '{mes_ano}' → substituindo")
            ws.cell(row=row_i, column=COL_DATA, value=mes_ano)
            matched += 1

    wb.save(XLSX_PATH)

    print(f"\n{'='*55}")
    print(f"✅ {matched} itens com DATA_LEITURA preenchida")
    if conflitos:
        print(f"⚠️  {conflitos} conflitos substituídos")
    print(f"❌ {len(nao_achado)} entradas não encontradas na coleção:\n")

    # Agrupa por mês para facilitar a leitura
    nao_achado.sort()
    mes_ant = None
    for mes_ano, titulo, edicao, score in nao_achado:
        if mes_ano != mes_ant:
            print(f"  [{mes_ano}]")
            mes_ant = mes_ano
        ed_str = f" #{edicao}" if edicao else ""
        conf   = f"  (match: {score:.0%})" if score > 0 else ""
        print(f"    • {titulo}{ed_str}{conf}")

    print(f"\n💾 Arquivo salvo: {XLSX_PATH}")
    print("\n⚠️  Lembre-se de NÃO abrir o quadrinhos.xlsx no Excel/Numbers enquanto o app roda.")


if __name__ == '__main__':
    main()
